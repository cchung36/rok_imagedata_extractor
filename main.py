import cv2
import os
from PIL import Image
from glob import glob
from openpyxl import Workbook
import numpy as np
import time

import pytesseract

cwd=os.path.dirname(os.path.realpath(__file__))

wb=Workbook()

sheet=wb.active
sheet.title="KD Power Ranking"
sheet['A1']="Kingdom #"
sheet['A2']="Top300 Power"
sheet['A3']="50M+ Player"
sheet['A4']="100M+ Player"

column=1

for root, subdirs, files in os.walk(cwd):

    for kd_num in subdirs:
        for File in os.listdir('.'+'/'+kd_num):
            if File.endswith(".PNG"):
                img_mask = kd_num+'/*.PNG'
                break
            elif File.endswith(".png"):
                img_mask = kd_num+'/*.png'
                break
            elif File.endswith(".JPG"):
                print(File)
                img_mask = kd_num+'/*.JPG'
                break
            elif File.endswith(".jpg"):
                img_mask = kd_num+'/*.jpg'
                break       

        img_names = glob(img_mask)        

        column=column+1

        sheet.cell(row=1, column=column).value = int(kd_num)

        powers = []

        for fn in img_names:

            print("Extracting data for image:{}".format(fn))

            img = cv2.imread(fn)

            if img.shape[0] == 1668 and img.shape[1] == 2224:
                crop_img = img[560:1380, 1650:1850]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"     
            elif img.shape[0] == 828 and img.shape[1] == 1792:
                crop_img = img[220:780, 1200:1400]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"
            elif img.shape[0] == 1125 and img.shape[1] == 2436:
                crop_img = img[300:1050, 1700:1900]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0 --psm 6"
            elif img.shape[0] == 1536 and img.shape[1] == 2048:
                crop_img = img[450:1280, 1500:1750]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"
            elif img.shape[0] == 1242 and img.shape[1] == 2688:
                crop_img = img[350:1150,1850:2100]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"
            elif img.shape[0] == 750 and img.shape[1] == 1334:
                crop_img = img[190:680, 980:1120]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"
            elif img.shape[0] == 1080 and img.shape[1] == 2340:
                crop_img = img[280:1000,1620:1820]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0 --psm 6"
            elif img.shape[0] == 817 and img.shape[1] == 1454:
                crop_img = img[200:730, 1050:1230]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"
            elif img.shape[0] == 900 and img.shape[1] == 1600:
                crop_img = img[250:850,1180:1350]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"
            elif img.shape[0] == 1080 and img.shape[1] == 2220:
                crop_img = img[280:1000,1550:1750]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0 --psm 6"
            elif img.shape[0] == 1080 and img.shape[1] == 2400:
                crop_img = img[280:1000,1650:1850]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0 --psm 6"
            elif img.shape[0] == 1440 and img.shape[1] == 2560:
                crop_img = img[400:1350, 1800:2150]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"
            elif img.shape[0] == 1242 and img.shape[1] == 2208:
                crop_img = img[350:1200,1600:1850]
                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0"


            gray = cv2.cvtColor(crop_img, cv2.COLOR_BGR2GRAY)

            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (crop_img.shape[1],1))

            blurred_img = cv2.GaussianBlur(crop_img,(5,5),0)

            kernel_sharpening = np.array([[-1,-1,-1], [-1, 9,-1],[-1,-1,-1]])

            sharpened_img = cv2.filter2D(blurred_img, -1, kernel_sharpening)

            gray = cv2.cvtColor(sharpened_img, cv2.COLOR_BGR2GRAY)

            _, thres_mask = cv2.threshold(gray,190,255,cv2.THRESH_BINARY) 
            _, line_mask = cv2.threshold(gray,160,255,cv2.THRESH_BINARY) 

#            cv2.imwrite('image.jpg', gray)
#            cv2.imwrite('image.jpg', thres_mask)

            line_mask = cv2.morphologyEx(line_mask, cv2.MORPH_OPEN,horizontal_kernel)
            cnts = cv2.findContours(line_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = cnts[0] if len(cnts) == 2 else cnts[1]

            for c in cnts:
                cv2.drawContours(thres_mask, [c], -1, 0, 3)

            nlabels, labels, stats, centroids = cv2.connectedComponentsWithStats(thres_mask, None, None, None, 8, cv2.CV_32S)
            sizes = stats[1:, -1] #get CC_STAT_AREA component
            thres_mask = np.zeros((labels.shape), np.uint8)

            for i in range(0, nlabels - 1):
                if sizes[i] >= 3:   #filter small dotted regions
                    thres_mask[labels == i + 1] = 255

            try:
                str_tokens=pytesseract.image_to_string(thres_mask,timeout=5,config=custom_config).split()
            except RuntimeError as timeout_error:
            # Tesseract processing is terminated
                pass
            
            if len(str_tokens) == 6:
                for item in str_tokens:
                    if item.replace(',','').isdigit():
                        power_value=int(item.replace(',',''))
                        if power_value >= 300000000 or power_value <= 15000000:
                            if custom_config == r"-c tessedit_char_whitelist=0123456789, --oem 0":
                                custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0 --psm 6"
                            else:
                                custom_config == r"-c tessedit_char_whitelist=0123456789, --oem 0"
                            break
                        else:
                            pass

                try:
                    str_tokens=pytesseract.image_to_string(thres_mask,timeout=5,config=custom_config).split()
                except RuntimeError as timeout_error:
                # Tesseract processing is terminated
                    pass

            else:
                if custom_config == r"-c tessedit_char_whitelist=0123456789, --oem 0":
                    custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0 --psm 6"
                else:
                    custom_config == r"-c tessedit_char_whitelist=0123456789, --oem 0"

                try:
                    str_tokens=pytesseract.image_to_string(thres_mask,timeout=5,config=custom_config).split()
                except RuntimeError as timeout_error:
                # Tesseract processing is terminated
                    pass

            print(str_tokens)

            for item in str_tokens:
                if item.replace(',','').isdigit():
                    powers.append(int(item.replace(',','')))

        power_arr=np.array(powers)

        power_arr[::-1].sort()

        print("Kingdom:{}".format(kd_num))
        print("player power:{}".format(power_arr))
        print("power size:{},total image numbers:{}".format(len(powers),len(img_names)))

        power_greater_than_50m=[item for item in powers if item > 50000000]
        power_greater_than_100m=[item for item in powers if item > 100000000]

#        total_power=np.sum(power_arr,dtype=np.int32)
        total_power=0

#        sheet.cell(row=2,column=column).value = total_power
        sheet.cell(row=3,column=column).value = len(power_greater_than_50m)
        sheet.cell(row=4,column=column).value = len(power_greater_than_100m)

        for i in range(len(powers)):
            sheet.cell(row=5+i, column=column).value = power_arr[i]
            total_power=total_power+power_arr[i]

        sheet.cell(row=2,column=column).value = total_power

        time.sleep(10)


wb.save('rok_data.xlsx')
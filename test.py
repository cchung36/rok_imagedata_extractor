import cv2
import os
from PIL import Image
from glob import glob
from openpyxl import Workbook
import openpyxl
import numpy as np
import time

import pytesseract

cwd=os.path.dirname(os.path.realpath(__file__))

wb=Workbook()

sample = 30

excel_path=cwd+"/KVK 2 MMR 20200925.xlsx"
ocr_path=cwd+"/rok_data.xlsx"

wb_obj = openpyxl.load_workbook(excel_path)
ocr_wb_obj = openpyxl.load_workbook(ocr_path)

sheet_obj = wb_obj.worksheets[1]
ocr_obj = ocr_wb_obj.active

max_col = sheet_obj.max_column 

for i in range(2,max_col):
    accuracy=0
    power=[]
    kd_num=sheet_obj.cell(row=1, column=i).value

    print("KD:{}".format(kd_num))

    for j in range(2,max_col):
        if kd_num == ocr_obj.cell(row=1, column=j).value:
            for k in range(5,sample+5):
                ocr_cell=ocr_obj.cell(row=k, column=j)
                excel_cell=sheet_obj.cell(row=k, column=i)
                print("excel value:{}, ocr value:{}".format(excel_cell.value,ocr_cell.value))
                if ocr_cell.value == excel_cell.value:
                    print("result match!")
                    accuracy=accuracy+1
                else:
                    print("result does not match!")
    
    print("accuracy:{}".format(accuracy/sample))

    time.sleep(20)
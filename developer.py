import cv2
import os
from PIL import Image
from glob import glob
from openpyxl import Workbook
import numpy as np

import pytesseract

cwd=os.path.dirname(os.path.realpath(__file__))

wb=Workbook()

sheet=wb.active
sheet.title="KD Power Ranking"
sheet['A1']="Kingdom #"
sheet['A2']="Top300 Power"
sheet['A3']="50M+ Player"
sheet['A4']="100M+ Player"

column=1

kd_num='1788'

for File in os.listdir('.'+'/'+kd_num):
    if File.endswith(".PNG"):
        img_mask = kd_num+'/*.PNG'
        break
    elif File.endswith(".png"):
        img_mask = kd_num+'/*.png'
        break
    elif File.endswith(".JPG"):
        print(File)
        img_mask = kd_num+'/*.JPG'
        break
    elif File.endswith(".jpg"):
        img_mask = kd_num+'/*.jpg'
        break

img_names = glob(img_mask)

powers = []

column=column+1

sheet.cell(row=1, column=column).value = int(kd_num)
num=1
for fn in img_names:
    img = cv2.imread(fn)
    (height,width,channel)=img.shape
    print("width:{},height:{}".format(width, height))

#    crop_img = img[560:1380, 1650:1850] #1813,1782,1810,1802 2224x1668
#    crop_img = img[220:780, 1200:1400] #1814,1801 1792x828
    crop_img = img[300:1050, 1700:1900] #1785,1788 1806 1784 1790 2436x1125
#    crop_img = img[450:1280, 1500:1750] #1807 2048x1536
#    crop_img = img[350:1150,1850:2100]#1798 2688x1242
#    crop_img = img[190:680, 980:1120] #1796,1804 1334x750
#    crop_img = img[280:1000,1620:1820] #1781 1823 1818 1787 2340x1080
#    crop_img = img[200:730, 1050:1230] #1824 1815 1454x817
#    crop_img = img[250:850,1180:1350] #1822 1820 1600x900
#    crop_img = img[280:1000,1550:1750] #1816 1777 2220x1080
#    crop_img = img[280:1000,1650:1850] #1811 1809 2400x1080
#    crop_img = img[180:620, 870:1020] #1829 1830 1198x673
#    crop_img = img[400:1350, 1800:2150] #1780 2560x1440
#    crop_img = img[350:1200,1600:1850] #1779 2208x1242

    gray = cv2.cvtColor(crop_img, cv2.COLOR_BGR2GRAY)
#    thresh = cv2.adaptiveThreshold(gray,255, cv2.ADAPTIVE_THRESH_MEAN_C,cv2.THRESH_BINARY,15,-2)
#    gray = cv2.cvtColor(blurred_img, cv2.COLOR_BGR2GRAY)

#    gray = cv2.equalizeHist(gray)

    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (crop_img.shape[1],1))
#    detected_lines= cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
#    cnts = cv2.findContours(detected_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
#    cnts = cnts[0] if len(cnts) == 2 else cnts[1]
#    for c in cnts:
#        cv2.drawContours(crop_img, [c], -1, (0,0,0), 2)

    blurred_img = cv2.GaussianBlur(crop_img,(5,5),0)

    kernel_sharpening = np.array([[-1,-1,-1],[-1, 9,-1],[-1,-1,-1]])

    sharpened_img = cv2.filter2D(blurred_img, -1, kernel_sharpening)


#    blur_img = cv2.GaussianBlur(crop_img,(5,5),0)

#    lab= cv2.cvtColor(sharpened_img, cv2.COLOR_BGR2LAB)
#    l, a, b = cv2.split(lab)
#    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
#    cl = clahe.apply(l)
#    limg = cv2.merge((cl,a,b))
#    clahe_img = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)

    gray = cv2.cvtColor(sharpened_img, cv2.COLOR_BGR2GRAY)
#    gray = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY) 
#    gray = cv2.GaussianBlur(gray,(5,5),0)    #1777

#    _, thres_mask = cv2.threshold(gray,242,255,cv2.THRESH_BINARY)
#    _, thres_mask = cv2.threshold(gray,230,255,cv2.THRESH_BINARY)
    _, thres_mask = cv2.threshold(gray,190,255,cv2.THRESH_BINARY) #1824
    _, line_mask = cv2.threshold(gray,160,255,cv2.THRESH_BINARY) #1777

    line_mask = cv2.morphologyEx(line_mask, cv2.MORPH_OPEN,horizontal_kernel)
    cnts = cv2.findContours(line_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]

    for c in cnts:
        cv2.drawContours(thres_mask, [c], -1, 0, 3)

    nlabels, labels, stats, centroids = cv2.connectedComponentsWithStats(thres_mask, None, None, None, 8, cv2.CV_32S)
    sizes = stats[1:, -1] #get CC_STAT_AREA component
    thres_mask = np.zeros((labels.shape), np.uint8)

    for i in range(0, nlabels - 1):
        if sizes[i] >= 3:   #filter small dotted regions
            thres_mask[labels == i + 1] = 255

#    thres_mask = cv2.bitwise_not(thres_mask)


#    thres_mask = cv2.erode(thres_mask,horizontal_kernel,iterations = 1)
#    thres_mask = cv2.dilate(thres_mask,horizontal_kernel,iterations = 1)


    cv2.imwrite('image.jpg', gray)
    cv2.imwrite('image_'+str(num)+'.jpg', thres_mask)

    num=num+1

#    custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0 --psm 6" #1813,1782,1785,1788
    custom_config=r"-c tessedit_char_whitelist=0123456789, --oem 0" #1814,1801,1796,1804 1824 1815
    str_tokens=pytesseract.image_to_string(thres_mask,timeout=5,config=custom_config).split()
#            str_tokens=pytesseract.image_to_string(gray,config=custom_config).split()
    print(str_tokens)

    for item in str_tokens:
        if item.replace(',','').isdigit():
            powers.append(int(item.replace(',','')))

power_arr=np.array(powers)

power_arr[::-1].sort()

print("Kingdom:{}".format(kd_num))
print("player power:{}".format(power_arr))
print("power size:{},total image numbers:{}".format(len(powers),len(img_names)))

power_greater_than_50m=[item for item in powers if item > 50000000]
power_greater_than_100m=[item for item in powers if item > 100000000]

total_power=np.sum(power_arr,dtype=np.int32)

sheet.cell(row=2,column=column).value = total_power
sheet.cell(row=3,column=column).value = len(power_greater_than_50m)
sheet.cell(row=4,column=column).value = len(power_greater_than_100m)

for i in range(len(powers)):
    sheet.cell(row=5+i, column=column).value = power_arr[i]


wb.save('rok_data.xlsx')

#    print(power_arr)
